import os
import openpyxl
from openpyxl import load_workbook
import sys
from win32com.client import Dispatch
from pathlib import Path
import re
from imdb_manager import IMDbData
from excel_manager import handle_excel_save
#from tmdbv3api import TMDb
#from tmdbv3api import Movie as tmdb_movie

input = str(" ".join(sys.argv[1:]))
if len(input) < 1:
    print("Try typing a movie name!")
    sys.exit(0)
print("Loading movie data...")

class Movie:
    def __init__(self, title, director, stars, plot, genre, imdb_rating, imdb_link, year, runtime, is_on_netflix):
        self.title = title
        self.director = director
        self.stars = stars
        self.plot = plot
        self.genre = genre
        self.imdb_rating = imdb_rating
        self.imdb_link = imdb_link
        self.year = year
        self.runtime = runtime
        self.is_on_netflix = is_on_netflix

class ExcelData:
    # The doc with netflix catalog
    doc_title = "netflix_titles"
    file_loc = "%s.xlsx" % doc_title
    workbook = openpyxl.load_workbook(file_loc)
    sheet = workbook[doc_title]
    title_column = 3

    # Our own watch list
    dir_path = str(Path.home())
    wb_path = os.path.join(dir_path, "temflix_movie_list.xlsx")

def get_input_without_args(user_input):
    word_list = user_input.split()
    if word_list[-1].lower() == "true" or word_list[-1].lower() == "false":
        user_input = user_input[:user_input.rfind(" ")]
    return user_input

def is_title_on_netflix(title):
    stripped_input = get_input_without_args(title)
    title_lowercase = stripped_input.lower()
    for i in range(1, ExcelData.sheet.max_row):
        excel_title_undercase = str(ExcelData.sheet.cell(row=i, column=ExcelData.title_column).value).lower()
        if (title_lowercase == excel_title_undercase):
            return True
    return False

#tmdb = TMDb()
#tmdb.api_key = os.getenv("TMB_API_KEY")
#tmdb_m = tmdb_movie()
#popular = tmdb_m.popular()
#for p in popular:
    #print(p.title)

imdb = IMDbData(get_input_without_args(input));
movie = Movie(imdb.title, imdb.director, imdb.stars, imdb.plot, imdb.genre, imdb.rating, imdb.url, imdb.year, imdb.runtime, is_title_on_netflix(input))

handle_excel_save(ExcelData.wb_path, movie, bool(re.search("true", input, re.IGNORECASE)))
