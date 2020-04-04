import openpyxl
from openpyxl import load_workbook
from win32com.client import Dispatch
import os

def handle_excel_save(wb_path, movie, should_open_excel):
    if os.path.exists(wb_path):
        wb = load_workbook(wb_path)
        add_new_title_to_watchlist(wb, movie, wb_path)
    else:
        wb = openpyxl.Workbook()
        initialize_watch_list(wb, wb_path)

    if should_open_excel:
        open_excel(wb_path)

def initialize_watch_list(wb, path):
    sheet = wb.active
    sheet.title = "Movie list"
    
    sheet["A1"] = "Title"
    sheet["B1"] = "Director"
    sheet["C1"] = "Stars"
    sheet["D1"] = "Plot"
    sheet["E1"] = "Genres"
    sheet["F1"] = "Rating"
    sheet["G1"] = "IMDb link"
    sheet["H1"] = "Release year"
    sheet["I1"] = "Runtime"
    sheet["J1"] = "On Netflix"
    
    wb.save(path)

def add_new_title_to_watchlist(wb, movie, path):
    sheet = wb.active
    empty_row = sheet.max_row + 1

    if (title_exists_in_watch_list(movie.title, sheet)):
        print("%s is already added to your watch list!" % movie.title)
        return

    sheet["A%s" % empty_row] = movie.title
    sheet["B%s" % empty_row] = movie.director
    sheet["C%s" % empty_row] = movie.stars
    sheet["D%s" % empty_row] = movie.plot
    sheet["E%s" % empty_row] = movie.genre
    sheet["F%s" % empty_row] = movie.imdb_rating
    sheet["G%s" % empty_row] = movie.imdb_link
    sheet["H%s" % empty_row] = movie.year
    sheet["I%s" % empty_row] = movie.runtime
    sheet["J%s" % empty_row] = movie.is_on_netflix

    wb.save(path)

def title_exists_in_watch_list(title_name, sheet):
    for i in range(1, sheet.max_row + 1):
        if title_name == str(sheet.cell(row=i, column=1).value):
            return True
    return False

def open_excel(path):
    xl = Dispatch("Excel.Application")
    xl.Visible = True # otherwise excel is hidden
    wb = xl.Workbooks.Open(path)
